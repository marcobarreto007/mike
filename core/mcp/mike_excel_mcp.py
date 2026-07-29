"""Safe local Excel workbook MCP server for Mike."""

from __future__ import annotations

import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries


PROJECT_ROOT = Path(
    os.getenv("MIKE_HOME") or Path(__file__).resolve().parents[2]
).resolve()
ROOT = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else PROJECT_ROOT
mcp = FastMCP("Mike Excel MCP", json_response=True)


def _resolve_path(raw_path: str, *, must_exist: bool = True) -> Path:
    candidate = Path(raw_path).expanduser()
    if not candidate.is_absolute():
        candidate = ROOT / candidate
    candidate = candidate.resolve()
    try:
        candidate.relative_to(ROOT)
    except ValueError as exc:
        raise ValueError(f"Path fora da raiz permitida: {ROOT}") from exc
    if candidate.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Somente arquivos .xlsx ou .xlsm sao permitidos")
    if must_exist and not candidate.exists():
        raise FileNotFoundError(str(candidate))
    return candidate


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _atomic_save(workbook, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.stem}.mike-tmp{target.suffix}")
    workbook.save(temporary)
    os.replace(temporary, target)


@mcp.tool(description="Lista planilhas Excel dentro da raiz autorizada do Mike.")
def list_spreadsheets(
    directory: str = ".",
    recursive: bool = True,
    limit: int = 100,
) -> list[dict[str, Any]]:
    base = Path(directory).expanduser()
    if not base.is_absolute():
        base = ROOT / base
    base = base.resolve()
    try:
        base.relative_to(ROOT)
    except ValueError as exc:
        raise ValueError(f"Diretorio fora da raiz permitida: {ROOT}") from exc
    pattern = "**/*.xls*" if recursive else "*.xls*"
    items = []
    for path in sorted(base.glob(pattern)):
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}:
            stat = path.stat()
            items.append({
                "path": str(path),
                "relative_path": str(path.relative_to(ROOT)),
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
            if len(items) >= max(1, min(int(limit), 500)):
                break
    return items


@mcp.tool(description="Inspeciona abas, dimensoes e propriedades de um workbook Excel.")
def inspect_workbook(path: str) -> dict[str, Any]:
    target = _resolve_path(path)
    workbook = load_workbook(target, read_only=True, data_only=False)
    try:
        sheets = [
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "dimension": sheet.calculate_dimension(),
            }
            for sheet in workbook.worksheets
        ]
        return {
            "path": str(target),
            "sheet_names": workbook.sheetnames,
            "sheets": sheets,
            "defined_names": [str(item) for item in workbook.defined_names.values()],
        }
    finally:
        workbook.close()


@mcp.tool(description="Le valores de uma aba Excel com limites seguros.")
def read_sheet(
    path: str,
    sheet_name: str = "",
    max_rows: int = 100,
    max_columns: int = 30,
    values_only: bool = True,
) -> dict[str, Any]:
    target = _resolve_path(path)
    workbook = load_workbook(target, read_only=True, data_only=values_only)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        row_limit = max(1, min(int(max_rows), 1000))
        col_limit = max(1, min(int(max_columns), 200))
        rows = [
            [_json_value(value) for value in row[:col_limit]]
            for row in sheet.iter_rows(max_row=row_limit, values_only=True)
        ]
        return {
            "path": str(target),
            "sheet": sheet.title,
            "rows": rows,
            "returned_rows": len(rows),
            "truncated": sheet.max_row > row_limit or sheet.max_column > col_limit,
        }
    finally:
        workbook.close()


@mcp.tool(description="Le um intervalo A1 especifico de uma aba Excel.")
def read_range(
    path: str,
    cell_range: str,
    sheet_name: str = "",
    values_only: bool = True,
) -> dict[str, Any]:
    target = _resolve_path(path)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if (max_row - min_row + 1) * (max_col - min_col + 1) > 20_000:
        raise ValueError("Intervalo excede o limite de 20.000 celulas")
    workbook = load_workbook(target, read_only=True, data_only=values_only)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = [
            [_json_value(cell.value) for cell in row]
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
        ]
        return {
            "path": str(target),
            "sheet": sheet.title,
            "range": cell_range,
            "values": rows,
        }
    finally:
        workbook.close()


@mcp.tool(description="Cria um novo workbook Excel sem sobrescrever arquivo existente.")
def create_workbook(
    path: str,
    sheet_name: str = "Sheet1",
    headers: Optional[list[Any]] = None,
) -> dict[str, Any]:
    target = _resolve_path(path, must_exist=False)
    if target.exists():
        raise FileExistsError(str(target))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name or "Sheet1"
    if headers:
        sheet.append(list(headers))
    _atomic_save(workbook, target)
    workbook.close()
    return {"created": True, "path": str(target), "sheet": sheet.title}


@mcp.tool(description="Escreve uma matriz de valores a partir de uma celula inicial.")
def write_cells(
    path: str,
    start_cell: str,
    values: list[list[Any]],
    sheet_name: str = "",
) -> dict[str, Any]:
    target = _resolve_path(path)
    if not values or not all(isinstance(row, list) for row in values):
        raise ValueError("values deve ser uma matriz nao vazia")
    if sum(len(row) for row in values) > 20_000:
        raise ValueError("Escrita excede o limite de 20.000 celulas")
    workbook = load_workbook(target)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        anchor = sheet[start_cell]
        for row_offset, row in enumerate(values):
            for col_offset, value in enumerate(row):
                sheet.cell(
                    row=anchor.row + row_offset,
                    column=anchor.column + col_offset,
                    value=value,
                )
        _atomic_save(workbook, target)
        return {
            "updated": True,
            "path": str(target),
            "sheet": sheet.title,
            "start_cell": start_cell,
            "cell_count": sum(len(row) for row in values),
        }
    finally:
        workbook.close()


@mcp.tool(description="Anexa linhas ao final de uma aba Excel.")
def append_rows(
    path: str,
    rows: list[list[Any]],
    sheet_name: str = "",
) -> dict[str, Any]:
    target = _resolve_path(path)
    if not rows or not all(isinstance(row, list) for row in rows):
        raise ValueError("rows deve ser uma matriz nao vazia")
    if sum(len(row) for row in rows) > 20_000:
        raise ValueError("Append excede o limite de 20.000 celulas")
    workbook = load_workbook(target)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        for row in rows:
            sheet.append(row)
        _atomic_save(workbook, target)
        return {
            "updated": True,
            "path": str(target),
            "sheet": sheet.title,
            "rows_appended": len(rows),
        }
    finally:
        workbook.close()


if __name__ == "__main__":
    mcp.run(transport="stdio")
