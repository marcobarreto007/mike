"""
Identity helpers, quick-reply builders, and dynamic system-prompt prefix.

Call configure() once at server startup to inject SOUL, PROJECT_ROOT, and
service-getter callables. All functions are safe to call before configure()
— they degrade gracefully (empty/None returns).

Locale and family data are loaded from config/identity.json at import time
(resolved via MIKE_HOME env var or a path relative to this file). The
hardcoded tuples below serve as fallbacks if the file is absent.
"""
from __future__ import annotations

import json
import logging
import os
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

_log = logging.getLogger("mike")


def _load_identity_config() -> Dict[str, Any]:
    """Load config/identity.json; returns {} if not found."""
    mike_home = os.getenv("MIKE_HOME")
    if mike_home:
        path = Path(mike_home) / "config" / "identity.json"
    else:
        path = Path(__file__).parent.parent.parent / "config" / "identity.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            _log.warning("Failed to load identity config from %s, using defaults", path)
    return {}


_IDENTITY_CFG: Dict[str, Any] = _load_identity_config()


# ---------------------------------------------------------------------------
# Injected state — populated by configure()
# ---------------------------------------------------------------------------
_soul: Dict[str, Any] = {}
_project_root: Path = Path(".")
_family_identify_profile_map: Dict[str, str] = _IDENTITY_CFG.get(
    "family_identify_profile_map",
    {
        "marco": "marco",
        "ana_paula": "anapaula",
        "rapha": "raphael",
        "alice": "alice",
        "matheus": "matheus",
    },
)

_get_monitor_fn: Callable = lambda: None
_get_consciousness_fn: Callable = lambda: None
_get_verifier_fn: Callable = lambda: None
_get_governance_fn: Callable = lambda: None
_get_autonomy_fn: Callable = lambda: None


def configure(
    soul: Dict[str, Any],
    project_root,
    *,
    get_monitor: Optional[Callable] = None,
    get_consciousness: Optional[Callable] = None,
    get_verifier: Optional[Callable] = None,
    get_governance: Optional[Callable] = None,
    get_autonomy: Optional[Callable] = None,
    family_identify_profile_map: Optional[Dict[str, str]] = None,
) -> None:
    global _soul, _project_root, _family_identify_profile_map
    global _get_monitor_fn, _get_consciousness_fn, _get_verifier_fn
    global _get_governance_fn, _get_autonomy_fn
    _soul = soul or {}
    _project_root = Path(project_root) if project_root else Path(".")
    if family_identify_profile_map is not None:
        _family_identify_profile_map = family_identify_profile_map
    if get_monitor is not None:
        _get_monitor_fn = get_monitor
    if get_consciousness is not None:
        _get_consciousness_fn = get_consciousness
    if get_verifier is not None:
        _get_verifier_fn = get_verifier
    if get_governance is not None:
        _get_governance_fn = get_governance
    if get_autonomy is not None:
        _get_autonomy_fn = get_autonomy


# ---------------------------------------------------------------------------
# Portuguese locale data — loaded from config/identity.json with hardcoded fallbacks
# ---------------------------------------------------------------------------
_PT_WEEKDAYS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("pt_weekdays", [
    "segunda-feira", "terca-feira", "quarta-feira",
    "quinta-feira", "sexta-feira", "sabado", "domingo",
]))
_PT_MONTHS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("pt_months", [
    "janeiro", "fevereiro", "marco", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]))
_DATE_QUERY_EXCLUSIONS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("date_query_exclusions", [
    "agenda", "clima", "compromisso", "email", "emails",
    "evento", "noticia", "noticias", "o que aconteceu",
    "previsao", "preco", "tarefa", "tarefas", "tempo hoje",
]))
_TIME_QUERY_EXCLUSIONS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("time_query_exclusions", [
    "distancia", "equacao", "problema", "raciocinio", "trem", "trens",
    "velocidade", "velocidades",
]))
_FAMILY_QUERY_MARKERS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("family_query_markers", [
    "pessoas que vc conhece", "pessoas que voce conhece", "todas as pessoas",
    "quem voce conhece", "quem vc conhece", "quem da familia",
    "o que vc se lembra", "o que voce se lembra", "memoria da familia",
    "soul",
]))
_FAMILY_INTENT_MARKERS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("family_intent_markers", [
    "quem", "quais", "lista", "listar", "me lembra", "se lembra", "recorda", "memoria", "soul",
]))
_LITERAL_REPLY_GUARDS: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("literal_reply_guards", [
    "responda exatamente", "repita exatamente", "apenas", "somente", "uma palavra",
]))
_DEFAULT_FAMILY_ORDER: Tuple[str, ...] = tuple(_IDENTITY_CFG.get("default_family_order", [
    "marco", "ana_paula", "rapha", "alice", "mike_dog",
    "marilene", "nilton_sulz", "best_friend", "marcelo_alves", "frederico_araujo",
]))
_FAMILY_DISPLAY_NAMES: Dict[str, str] = _IDENTITY_CFG.get("family_display_names", {
    "ana_paula": "Ana Paula",
    "rapha": "Rapha",
    "mike_dog": "Mike, o yorkshire",
    "marilene": "Marilene",
    "nilton_sulz": "Nilton Sulz",
    "best_friend": "Melhor amigo do Marco",
    "marcelo_alves": "Marcelo Alves",
    "frederico_araujo": "Frederico Araujo",
})


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------

def normalize_identity_text(text: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(text or ""))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def family_display_name(key: str, info: dict) -> str:
    if info.get("full_name"):
        return str(info["full_name"])
    return _FAMILY_DISPLAY_NAMES.get(key, key.replace("_", " ").title())


def compact_soul_person_detail(info: dict) -> str:
    parts = []
    relation = str(info.get("relation") or "").strip()
    if relation:
        parts.append(relation)
    for field in ("work", "studies", "grades", "school", "personality", "dream", "note"):
        value = info.get(field)
        if value:
            parts.append(str(value).strip())
    interests = info.get("interests")
    if isinstance(interests, list) and interests:
        parts.append("Interesses: " + ", ".join(str(item) for item in interests[:6]))
    detail = ". ".join(part for part in parts if part)
    if len(detail) > 280:
        detail = detail[:277].rstrip() + "..."
    return detail or "Pessoa importante na memoria familiar do Mike."


# ---------------------------------------------------------------------------
# Quick-reply builders
# ---------------------------------------------------------------------------
# Natural identification — sem passwords
# ---------------------------------------------------------------------------

def format_natural_identification(user_text: str, profile_key: Optional[str] = None) -> Optional[str]:
    """Detect when someone identifies themselves and respond naturally.

    No passwords. No login screens. Just conversation.
    """
    if not user_text:
        return None
    try:
        from mike_auth import detect_profile_from_message
    except ImportError:
        return None
    detected = detect_profile_from_message(user_text)
    if not detected:
        return None
    # Map profile keys to friendly names
    names = {
        "marco": "Marco", "anapaula": "Ana Paula",
        "raphael": "Raphael", "alice": "Alice",
        "matheus": "Matheus", "marilene": "Marilene",
        "visitante": "Visitante",
    }
    name = names.get(detected, detected.title())
    greetings = {
        "marco": f"Fala, Chefe! 🐾 Soube na hora que era você. O que precisa?",
        "anapaula": f"Ana Paula! 🐾 Rainha da casa! Como posso ajudar?",
        "raphael": f"Rapha! 🐾 Meu garoto! Tudo bem contigo?",
        "alice": f"Alice! 🐾 Princesa! Que bom falar contigo!",
        "matheus": f"Matheus! 🐾 Bem-vindo! Em que posso ajudar?",
        "marilene": f"Marilene! 🐾 Que bom ter você aqui! Como posso ajudar?",
    }
    greeting = greetings.get(detected, f"Olá, {name}! 🐾 Que bom falar contigo! Em que posso ajudar?")
    return greeting


# ---------------------------------------------------------------------------

def format_current_datetime_reply(user_text: str) -> Optional[str]:
    normalized = normalize_identity_text(user_text)
    if not normalized:
        return None
    if any(marker in normalized for marker in _DATE_QUERY_EXCLUSIONS):
        return None
    if any(marker in normalized for marker in _TIME_QUERY_EXCLUSIONS):
        return None

    asks_today = "hoje" in normalized and (
        "que dia" in normalized
        or "qual dia" in normalized
        or "data" in normalized
        or "dia e hoje" in normalized
        or "dia eh hoje" in normalized
    )
    asks_time = any(
        marker in normalized
        for marker in ("que horas", "qual hora", "hora atual", "horario atual")
    )
    if not asks_today and not asks_time:
        return None

    direct_time_query = bool(
        re.match(r"^(mike\s+)?(que horas|qual hora|hora atual|horario atual)\b", normalized)
    )
    if asks_time and not asks_today and not direct_time_query:
        return None

    now_dt = datetime.now()
    weekday = _PT_WEEKDAYS[now_dt.weekday()]
    month = _PT_MONTHS[now_dt.month - 1]
    date_text = f"{now_dt.day} de {month} de {now_dt.year}"

    if asks_time and not asks_today:
        return f"Agora sao {now_dt.strftime('%H:%M')}."
    if asks_time:
        return f"Hoje e {weekday}, {date_text}. Agora sao {now_dt.strftime('%H:%M')}."
    return f"Hoje e {weekday}, {date_text}."


def format_family_memory_reply(user_text: str) -> Optional[str]:
    normalized = normalize_identity_text(user_text)
    if not normalized:
        return None
    if any(guard in normalized for guard in _LITERAL_REPLY_GUARDS):
        return None
    has_family_marker = any(marker in normalized for marker in _FAMILY_QUERY_MARKERS)
    has_intent_marker = any(marker in normalized for marker in _FAMILY_INTENT_MARKERS)
    if not (has_family_marker and has_intent_marker):
        return None

    family = _soul.get("family", {}) if isinstance(_soul, dict) else {}
    if not isinstance(family, dict) or not family:
        return "Marco, eu nao encontrei a Soul carregada agora. Preciso revisar o arquivo de memoria do Mike."

    ordered_keys = [key for key in _DEFAULT_FAMILY_ORDER if isinstance(family.get(key), dict)]
    ordered_keys.extend(
        key for key, value in family.items()
        if key not in ordered_keys and isinstance(value, dict)
    )

    lines = ["Marco, eu tenho a Soul carregada sim. Da nossa familia, eu tenho salvo:"]
    for key in ordered_keys:
        info = family.get(key) or {}
        display = family_display_name(key, info)
        lines.append(f"- {display}: {compact_soul_person_detail(info)}")
    return "\n".join(lines).strip()


def maybe_create_default_spreadsheet_reply(user_text: str) -> Optional[str]:
    normalized = normalize_identity_text(user_text)
    if not normalized:
        return None
    mentions_sheet = any(term in normalized for term in ("planilha", "planilia", "excel", "xcel", "xlsx"))
    create_intent = any(
        term in normalized
        for term in ("faca", "faz", "crie", "criar", "monte", "monta", "gere", "gerar")
    )
    if not mentions_sheet or not create_intent:
        return None

    target_dir = _project_root / "data" / "planilhas"
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "planilha_modelo_mike.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Controle"
        headers = ["Data", "Categoria", "Descricao", "Valor", "Responsavel", "Status", "Observacoes"]
        ws.append(headers)
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        rows = [
            [today.isoformat(), "Entrada", "Receita exemplo", 2500, "Marco", "Confirmado", ""],
            [today.isoformat(), "Saida", "Despesa exemplo", -350, "Marco", "Pendente", ""],
            [tomorrow.isoformat(), "Tarefa", "Ligar para cliente", 0, "Marco", "Aberto", "Follow-up"],
        ]
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for idx, width in enumerate([14, 16, 28, 14, 18, 16, 32], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

        summary = wb.create_sheet("Resumo")
        summary["A1"] = "Indicador"
        summary["B1"] = "Valor"
        for cell in summary[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
        summary.append(["Total de entradas", '=SUMIF(Controle!B:B,"Entrada",Controle!D:D)'])
        summary.append(["Total de saidas", '=SUMIF(Controle!B:B,"Saida",Controle!D:D)'])
        summary.append(["Saldo", "=B2+B3"])
        summary.append(["Itens em aberto", '=COUNTIF(Controle!F:F,"Aberto")+COUNTIF(Controle!F:F,"Pendente")'])
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 18
        wb.save(target)
    except Exception as exc:
        return f"Marco, tentei criar a planilha Excel, mas falhou: {type(exc).__name__}: {exc}"

    return (
        "Pronto, Marco. Criei uma planilha Excel modelo com abas de Controle e Resumo, "
        f"incluindo formulas basicas de total, saldo e itens em aberto: {target}"
    )


def maybe_builtin_chat_reply(user_text: str, profile_key: Optional[str] = None) -> Optional[str]:
    del profile_key  # reserved for per-profile rules in future
    for responder in (
        format_self_check_reply,         # must run first — real file I/O diagnostics
        format_natural_identification,   # "sou o Marco" → "Fala, Marco!"
        format_current_datetime_reply,
        format_family_memory_reply,
    ):
        reply = responder(user_text)
        if reply:
            return reply
    return None


# ---------------------------------------------------------------------------
# Identity matching
# ---------------------------------------------------------------------------

def identity_aliases(profile_key: str, info: dict) -> Tuple[str, ...]:
    dash_profile = _family_identify_profile_map.get(profile_key, profile_key)
    aliases: set[str] = set()
    raw_candidates: List[Any] = [
        profile_key,
        profile_key.replace("_", " "),
        dash_profile,
        dash_profile.replace("_", " "),
        info.get("keyword"),
        info.get("full_name"),
    ]
    extra_aliases = info.get("aliases")
    if isinstance(extra_aliases, list):
        raw_candidates.extend(extra_aliases)

    for candidate in raw_candidates:
        normalized = normalize_identity_text(candidate)
        if not normalized:
            continue
        aliases.add(normalized)
        compact = normalized.replace(" ", "")
        if compact and compact != normalized:
            aliases.add(compact)

    return tuple(sorted(aliases, key=lambda item: (-len(item), item)))


def match_dashboard_profile_from_identity(text: str) -> Optional[str]:
    normalized_text = normalize_identity_text(text)
    if not normalized_text:
        return None

    for profile_key, info in _soul.get("family", {}).items():
        if not isinstance(info, dict):
            continue
        for alias in identity_aliases(profile_key, info):
            if normalized_text == alias:
                return _family_identify_profile_map.get(profile_key, profile_key)
            if re.search(rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])", normalized_text):
                return _family_identify_profile_map.get(profile_key, profile_key)

    return None


# ---------------------------------------------------------------------------
# Dynamic system-prompt prefix
# ---------------------------------------------------------------------------

def build_dynamic_prefix(user_text: str = "") -> str:
    """Prefixo dinamico por pedido: data/hora atual, alertas ativos, restricoes de projeto."""
    now_dt = datetime.now()
    dias = ["Segunda", "Terca", "Quarta", "Quinta", "Sexta", "Sabado", "Domingo"]
    day_name = dias[now_dt.weekday()]
    now_str = f"{day_name}, {now_dt.strftime('%d/%m/%Y %H:%M')}"
    lines = [
        "[CONTEXTO DO SISTEMA]",
        f"Hoje e: {now_str}",
        "USE ESTA DATA se perguntarem que dia ou hora e.",
    ]
    try:
        mon = _get_monitor_fn()
        if mon is not None:
            snap = mon.status()
            alerts = snap.get("alerts", [])
            if alerts:
                msgs = "; ".join(a.get("message", "") for a in alerts[:3])
                lines.append(f"Alertas ativos: {msgs}")
    except Exception:
        _log.exception("Failed to fetch monitor alerts for system context")

    if user_text:
        try:
            consciousness = _get_consciousness_fn()
            if consciousness:
                constraint_block = consciousness.build_constraint_block(user_text, max_constraints=3)
                if constraint_block:
                    lines.append(constraint_block)
        except Exception:
            _log.exception("Failed to build consciousness constraint block")

        try:
            verifier = _get_verifier_fn()
            if verifier:
                consciousness = _get_consciousness_fn()
                constraints = consciousness.get_constraints(user_text) if consciousness else []
                if constraints:
                    verification_prompt = verifier.inject_verification_prompt(constraints[:3])
                    if verification_prompt:
                        lines.append(verification_prompt)
        except Exception:
            _log.exception("Failed to inject verification prompt")

    try:
        gov = _get_governance_fn()
        if gov and gov._last_cycle and not gov._last_cycle.get("healthy", True):
            lines.append("[GOVERNANCE] Sistema com problemas — governanca ativa")
    except Exception:
        _log.exception("Failed to check governance status")

    try:
        autonomy = _get_autonomy_fn()
        if autonomy and autonomy._loaded:
            st = autonomy.status()
            tasks = st.get("tasks", {})
            emails = st.get("email_tracking", {})
            parts = []
            if tasks.get("pending", 0) > 0:
                parts.append(f"{tasks['pending']} tarefa(s) pendente(s) na Lousa")
            if tasks.get("running", 0) > 0:
                parts.append(f"{tasks['running']} tarefa(s) executando")
            if emails.get("overdue", 0) > 0:
                parts.append(f"{emails['overdue']} email(s) sem resposta (overdue)")
            if emails.get("waiting", 0) > 0:
                parts.append(f"{emails['waiting']} email(s) aguardando resposta")
            if parts:
                lines.append("[LOUSA] " + "; ".join(parts))
            elif tasks.get("done", 0) > 0:
                lines.append(f"[LOUSA] Tudo em dia! {tasks['done']} tarefa(s) concluida(s).")
    except Exception:
        _log.exception("Failed to fetch autonomy status")

    # -------------------------------------------------------------------
    # Self-check: handled by format_self_check_reply()
    # -------------------------------------------------------------------

    # -------------------------------------------------------------------
    # Natural identification: sem passwords, Mike identifica por nome
    # -------------------------------------------------------------------
    if user_text:
        try:
            from mike_auth import detect_profile_from_message
            detected = detect_profile_from_message(user_text)
            if detected:
                lines.append(f"[PERFIL] Voce esta falando com {detected.upper()}. Trate-o como tal — use o nome dele(a), o tom certo, e as preferencias do perfil.")
        except ImportError:
            pass

    # -------------------------------------------------------------------
    # Web search enforcement: for factual/current queries, FORCE web search
    # -------------------------------------------------------------------
    if _requires_web_search(user_text):
        lines.append(
            "[PESQUISA WEB OBRIGATORIA] O usuario esta a pedir informacao factual, "
            "atual ou especifica que voce NAO pode saber de cor (meteo, noticias, precos, "
            "eventos, dados atuais). USE AS FERRAMENTAS DE BUSCA AGORA: "
            "tavily_search ou puppeteer_navigate para pesquisar na web. "
            "NAO invente. NAO use 'conhecimento interno'. "
            "Se a busca falhar, diga HONESTAMENTE que nao conseguiu pesquisar."
        )

    return "\n".join(lines)


# -----------------------------------------------------------------------
# Self-check: real system diagnostics (reads actual files, not LLM guess)
# -----------------------------------------------------------------------

_SELF_CHECK_PATTERNS = [
    r"\breavalie\b.*\b(sistema|seu|proprio)\b",
    r"\bauto[-\s]?(avalie|verifique|diagnostico|check)\b",
    r"\bdiagnostico\b.*\b(sistema|interno|completo)\b",
    r"\b(verifique|verifica|cheque|checa)\b.*\b(seu|proprio|teu)\b.*\b(sistema|estado|funcionamento)\b",
    r"\bcomo\b.*\b(esta|estao)\b.*\b(funcionando|sistema|servidor|maquina)\b",
    r"\b(faca|faz|rode|executa)\b.*\b(auto[-\s]?(avaliacao|diagnostico|check|verificacao))\b",
    r"\b(self[-\s]?(check|diagnos|verif|eval|assess))\b",
    r"\bstatus\b.*\b(completo|geral|total|sistema|servidor)\b",
]


def _is_self_check_request(user_text: str) -> bool:
    """Return True if user_text looks like a self-check/self-evaluation request."""
    if not user_text:
        return False
    text_lower = user_text.lower().strip()
    for pattern in _SELF_CHECK_PATTERNS:
        if re.search(pattern, text_lower):
            return True
    return False


# -----------------------------------------------------------------------
# Web search patterns — queries that REQUIRE internet search
# -----------------------------------------------------------------------

_WEB_SEARCH_REQUIRED_PATTERNS = [
    # Weather / meteorology
    r"\b(meteo|meteorologia|previsao|previsão|tempo|clima|temperatura|chuva|neve|vento|umidade|umidade|graus|celsius|fahrenheit)\b",
    # News / current events
    r"\b(noticia|notícia|noticias|notícias|aconteceu|hoje|hoje|ultima|última|recente|jornal|atual|atual)\b",
    # Prices / stocks / finance
    r"\b(cotacao|cotação|preco|preço|bolsa|acao|ação|dolar|dólar|euro|bitcoin|ibovespa|nasdaq|sp500)\b",
    # Sports / games
    r"\b(jogo|placar|resultado|campeonato|futebol|basquete|tenis|tênis|ganhou|perdeu|empate)\b",
    # Factual lookups
    r"\b(pesquisar|pesquisa|pesquise|procura|busca|buscar|busque|procure|procura|ache|acha|encontra|encontre)\b",
    # Specific facts
    r"\b(quem (ganhou|venceu|elegeu|foi)|o que (aconteceu|houve|se passou)|qual (o|a) (valor|preco|preço|resultado))\b",
    # "search for X" / "look up X"
    r"\b(search|look.up|find|get.*(info|information|data))\b",
]


def _requires_web_search(user_text: str) -> bool:
    """Return True if the query likely requires internet search for factual data."""
    if not user_text:
        return False
    text_lower = user_text.lower().strip()
    for pattern in _WEB_SEARCH_REQUIRED_PATTERNS:
        if re.search(pattern, text_lower):
            return True
    return False


def format_self_check_reply(user_text: str, profile_key: Optional[str] = None) -> Optional[str]:
    """Real system self-diagnostics.

    Reads actual log files, soul.json, and health endpoint to produce
    a concrete diagnostic report — no LLM hallucination.
    """
    if not _is_self_check_request(user_text):
        return None

    import urllib.request as _urllib

    lines: list[str] = []
    lines.append("📊 **DIAGNÓSTICO REAL DO SISTEMA MIKE**\n")

    # ── 1. Server health (internal check, avoids deadlock of calling self) ──
    lines.append("### 🟢 Servidor")
    try:
        import time as _time
        import psutil as _psutil
        proc = _psutil.Process()
        uptime_s = _time.time() - proc.create_time()
        mem_mb = proc.memory_info().rss / (1024 * 1024)
        cpu_pct = proc.cpu_percent(interval=0.1)
        lines.append(f"- PID: {proc.pid}")
        lines.append(f"- Uptime: {uptime_s:.0f}s ({uptime_s/60:.1f} min)")
        lines.append(f"- Memoria: {mem_mb:.0f} MB")
        lines.append(f"- CPU: {cpu_pct:.1f}%")
        # Check if Qwen llama-server is reachable
        try:
            qwen_req = _urllib.Request("http://127.0.0.1:8081/health", method="GET")
            with _urllib.urlopen(qwen_req, timeout=3) as q_resp:
                if q_resp.status == 200:
                    lines.append("- LLM (Qwen 3.6): ✅ online (porta 8081)")
                else:
                    lines.append(f"- LLM (Qwen 3.6): ⚠️ status {q_resp.status}")
        except Exception:
            lines.append("- LLM (Qwen 3.6): ❌ nao respondeu (porta 8081)")
    except ImportError:
        lines.append("- psutil nao disponivel — sem metricas de processo")
    except Exception as exc:
        lines.append(f"- Erro ao obter metricas: {exc}")

    # ── 2. Log files ──
    def _tail(path: Path, n: int = 20) -> list[str]:
        try:
            if not path.exists():
                return [f"  (arquivo nao encontrado: {path})"]
            text = path.read_text(encoding="utf-8", errors="replace")
            all_lines = text.strip().splitlines()
            return all_lines[-n:]
        except Exception as exc:
            return [f"  (erro ao ler: {exc})"]

    log_dir = _project_root / "logs"

    lines.append("\n### 📋 Últimas 15 linhas de mike.log")
    for log_line in _tail(log_dir / "mike.log", 15):
        lines.append(f"  {log_line}")

    lines.append("\n### ⚠️ Últimas 10 linhas de mike_stderr.log")
    for err_line in _tail(log_dir / "mike_stderr.log", 10):
        lines.append(f"  {err_line}")

    # ── 3. Soul check ──
    soul_path = _project_root / "runtime" / "memory" / "soul.json"
    lines.append(f"\n### 🧠 Alma (soul.json)")
    if soul_path.exists():
        try:
            soul = json.loads(soul_path.read_text(encoding="utf-8"))
            family = soul.get("family", {})
            lines.append(f"- Membros da familia: {len(family)}")
            for name in list(family.keys())[:5]:
                lines.append(f"  - {name}")
            identity = soul.get("identity", {})
            if identity:
                lines.append(f"- Identidade: {identity.get('name', '?')} — {identity.get('role', '?')}")
        except Exception as exc:
            lines.append(f"- ERRO ao ler: {exc}")
    else:
        lines.append(f"- ❌ Arquivo nao encontrado: {soul_path}")

    # ── 4. Git status ──
    import subprocess as _sp
    lines.append("\n### 🔧 Git Status")
    try:
        result = _sp.run(
            ["git", "log", "--oneline", "-5"],
            cwd=str(_project_root),
            capture_output=True,
            text=True,
            timeout=10,
        )
        if result.returncode == 0:
            for commit_line in result.stdout.strip().splitlines():
                lines.append(f"  {commit_line}")
        else:
            lines.append(f"  git log falhou: {result.stderr.strip()}")
    except Exception as exc:
        lines.append(f"  git indisponivel: {exc}")

    # ── 5. Disk/RAM ──
    lines.append("\n### 💾 Recursos")
    try:
        import shutil as _shutil
        usage = _shutil.disk_usage(str(_project_root))
        free_gb = usage.free / (1024**3)
        total_gb = usage.total / (1024**3)
        pct = (free_gb / total_gb) * 100
        lines.append(f"- Disco: {free_gb:.1f} GB livre de {total_gb:.1f} GB ({pct:.0f}%)")
    except Exception as exc:
        lines.append(f"- Disco: nao foi possivel verificar — {exc}")

    lines.append(f"\n---\n*Diagnostico gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — dados REAIS do sistema.*")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Back-compat aliases (used in mike_server.py via _ prefix convention)
# ---------------------------------------------------------------------------
_normalize_identity_text = normalize_identity_text
_family_display_name = family_display_name
_compact_soul_person_detail = compact_soul_person_detail
_format_current_datetime_reply = format_current_datetime_reply
_format_family_memory_reply = format_family_memory_reply
_maybe_create_default_spreadsheet_reply = maybe_create_default_spreadsheet_reply
_maybe_builtin_chat_reply = maybe_builtin_chat_reply
_identity_aliases = identity_aliases
_match_dashboard_profile_from_identity = match_dashboard_profile_from_identity
_build_dynamic_prefix = build_dynamic_prefix
